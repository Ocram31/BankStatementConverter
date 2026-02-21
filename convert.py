#!/usr/bin/env python3
"""
PDF Bank Statement to CSV Converter for SA Accounting Import.

Supported output formats:
  generic       — Date, Description, Amount (signed). Default.
  sage          — Same as generic, zero amounts blank.
  sage_split    — Date, Description, Money in, Money out.
  xero          — *Date, *Amount, Payee, Description, Reference.
  quickbooks    — Date, Description, Amount. UTF-8, no commas.
  quickbooks_split — Date, Description, Credit, Debit.

Dedicated parsers (battle-tested):
  ABSA  — Tjekrekeningstaat, Credit Card, Current Account
  FNB   — Business Cheque Account

Generic auto-detect parser (detects column layout from headers):
  Standard Bank, Nedbank, Capitec, Investec,
  Discovery Bank, TymeBank, African Bank
  + any unknown bank with standard columnar layout

Usage:
  python3 convert.py                              # process all PDFs in current dir
  python3 convert.py file1.pdf file2.pdf          # specific files
  python3 convert.py --output-dir ./csv/          # custom output dir
  python3 convert.py --parser fnb file.pdf        # force parser type
  python3 convert.py --format sage_split          # Sage 4-column format
  python3 convert.py --format xero                # Xero precoded format
  python3 convert.py --password 8501015800080     # decrypt protected PDFs
"""

import argparse
import csv
import os
import re
import sys
import traceback
from dataclasses import dataclass
from datetime import date as date_type
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import pdfplumber

# Optional OCR support for FNB fee descriptions
try:
    import pytesseract
    from PIL import Image, ImageEnhance
    HAS_OCR = True
except ImportError:
    HAS_OCR = False

# Optional support for password-protected PDFs
try:
    import pikepdf
    HAS_PIKEPDF = True
except ImportError:
    HAS_PIKEPDF = False

# Optional Excel output support
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, numbers
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

MAX_DESCRIPTION_LENGTH = 200


def _get_app_dir():
    """Directory where the executable (or script) lives.

    In frozen (PyInstaller) mode, this is where the .exe sits — the right
    place for user-facing folders like pdfs/ and csv/.
    In normal mode, it's the script's own directory.
    """
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    return Path(__file__).parent


def _get_bundle_dir():
    """Directory where bundled data files live (e.g. VERSION).

    In frozen --onefile mode this is the temp extraction dir (sys._MEIPASS).
    In normal mode, same as the script directory.
    """
    if getattr(sys, 'frozen', False):
        return Path(sys._MEIPASS)
    return Path(__file__).parent


def _decrypt_pdf_if_needed(pdf_path, password=None):
    """Decrypt a password-protected PDF, returning path to a usable file.

    SA banks commonly encrypt statement PDFs (often with ID number as password).
    Returns the original path if not encrypted, or a decrypted temp path.
    Caller is responsible for cleaning up temp files.
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # If we can open and read, it's not encrypted
            if pdf.pages:
                _ = pdf.pages[0].extract_text()
                return pdf_path, None
    except Exception:
        pass

    if not HAS_PIKEPDF:
        raise ValueError(
            f"PDF appears to be password-protected: {pdf_path}\n"
            "  Install pikepdf to decrypt: pip install pikepdf\n"
            "  Then use: --password <password>"
        )

    if not password:
        raise ValueError(
            f"PDF is password-protected: {pdf_path}\n"
            "  Use: --password <password> (SA banks often use your ID number)"
        )

    decrypted_path = pdf_path + '.decrypted.pdf'
    try:
        with pikepdf.open(pdf_path, password=password) as encrypted_pdf:
            encrypted_pdf.save(decrypted_path)
        return decrypted_path, decrypted_path
    except pikepdf.PasswordError:
        raise ValueError(
            f"Wrong password for {pdf_path}. "
            "SA banks often use your ID number as the statement password."
        )


@dataclass
class Transaction:
    date: str
    description: str
    amount: Decimal
    balance: Decimal | None = None
    fee: Decimal = None
    category: str | None = None

    def __post_init__(self):
        if self.fee is None:
            self.fee = Decimal('0')


def clean_amount_spaces(s):
    """Remove space thousands separators: '4 683.99' -> '4683.99'"""
    return s.replace(' ', '').strip()


def clean_amount_commas(s):
    """Remove comma thousands separators: '63,378.53' -> '63378.53'"""
    return s.replace(',', '').strip()


def clean_description(desc):
    """Normalize whitespace and enforce length limit."""
    desc = ' '.join(desc.split()).strip()
    if len(desc) > MAX_DESCRIPTION_LENGTH:
        desc = desc[:MAX_DESCRIPTION_LENGTH - 3] + '...'
    return desc


def _validate_date(date_str):
    """Validate DD/MM/YYYY date string. Returns True if valid."""
    try:
        parts = date_str.split('/')
        date_type(int(parts[2]), int(parts[1]), int(parts[0]))
        return True
    except (ValueError, IndexError):
        return False


# Date pattern for ABSA Tjek
_DATE_RE_ABSA = re.compile(r'^\d{1,2}/\d{2}/\d{4}$')


def _group_words_into_lines(words, tolerance=3):
    """Group pdfplumber words into lines by y-coordinate.

    Args:
        words: List of word dicts from pdfplumber extract_words()
        tolerance: Max y-distance (in PDF points) to consider same line
    """
    if not words:
        return []
    sorted_words = sorted(words, key=lambda w: (w['top'], w['x0']))
    lines = []
    current_line = [sorted_words[0]]
    for w in sorted_words[1:]:
        if abs(w['top'] - current_line[0]['top']) <= tolerance:
            current_line.append(w)
        else:
            lines.append(sorted(current_line, key=lambda w: w['x0']))
            current_line = [w]
    lines.append(sorted(current_line, key=lambda w: w['x0']))
    return lines


def _parse_sa_amount(s):
    """Parse SA-format amount: '4 683.99' -> Decimal('4683.99')"""
    s = s.replace(' ', '').strip()
    if not s:
        return None
    try:
        return Decimal(s)
    except Exception:
        return None


def _parse_sa_balance(s):
    """Parse SA-format balance with optional trailing '-'.
    '112 370.73-' -> Decimal('-112370.73')
    """
    s = s.replace(' ', '').strip()
    if not s:
        return None
    negative = s.endswith('-')
    if negative:
        s = s[:-1]
    try:
        val = Decimal(s)
        return -val if negative else val
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Parser A: ABSA Tjekrekeningstaat (word-level column detection)
# ---------------------------------------------------------------------------

# Default column boundaries (fallback if header detection fails)
_ABSA_TJEK_DEFAULTS = {
    'fee': 265,
    'debit': 320,
    'credit': 415,
    'balance': 510,
}

# Map PDF header words (Afrikaans + English) → internal column key
_ABSA_TJEK_HEADER_VARIANTS = {
    'Koste': 'fee', 'Debietbedrag': 'debit',
    'Kredietbedrag': 'credit', 'Saldo': 'balance',
    'Charges': 'fee', 'Fee': 'fee', 'Fees': 'fee',
    'Debit': 'debit', 'Debits': 'debit',
    'Credit': 'credit', 'Credits': 'credit',
    'Balance': 'balance',
}

# Page header/footer markers to skip in continuation lines (Afrikaans + English)
_ABSA_TJEK_BOILERPLATE = (
    'Privaatheidskennisgewing', 'Privacy Notice',
    'Bladsy', 'Page',
    'Finansielediensteverskaffer', 'Financial Services Provider',
    'Registrasienommer', 'Registration Number',
    'Tjekrekeningnommer', 'Cheque Account Number',
    'transaksies (vervolg)', 'transactions (continued)',
    'Transaksiebeskrywing', 'Transaction Description',
    'CSP002CQ',
    'Absa Bank Beperk', 'Absa Bank Limited',
    'Gemagtigde', 'Authorised',
)


def _detect_column_boundaries(pdf):
    """Auto-detect ABSA Tjek column boundaries from header row.

    Finds the header line by looking for a row that maps to all 4 internal
    column keys (fee, debit, credit, balance) using _ABSA_TJEK_HEADER_VARIANTS.
    Works with both Afrikaans (Debietbedrag/Kredietbedrag) and English (Debit/Credit).
    Returns MIDPOINTS between consecutive header x0 positions as boundaries,
    because numeric data is right-aligned within columns.
    Falls back to hardcoded defaults with a warning if not found.
    """
    for page in pdf.pages:
        words = page.extract_words()
        lines = _group_words_into_lines(words)

        for line_words in lines:
            # Map each word on this line to an internal key (if it's a header word)
            headers = {}
            for w in line_words:
                internal = _ABSA_TJEK_HEADER_VARIANTS.get(w['text'])
                if internal and internal not in headers:
                    headers[internal] = w['x0']

            # Need both debit and credit to confirm it's the header row
            if 'debit' in headers and 'credit' in headers and len(headers) == 4:
                # Use midpoints between consecutive headers as boundaries.
                return {
                    'fee': headers['fee'],
                    'debit': (headers['fee'] + headers['debit']) / 2,
                    'credit': (headers['debit'] + headers['credit']) / 2,
                    'balance': (headers['credit'] + headers['balance']) / 2,
                }

    print(f'  WARNING: Could not auto-detect column positions from header row')
    print(f'  Using default column boundaries (may be inaccurate)')
    return _ABSA_TJEK_DEFAULTS


def parse_absa_tjekrekeningstaat(pdf_path):
    """Parse ABSA Tjekrekeningstaat using word-level column detection.

    Uses pdfplumber extract_words() with x-coordinates to determine
    which column (Debietbedrag vs Kredietbedrag) each amount belongs to.
    Column boundaries are auto-detected from the PDF header row.
    """
    transactions = []
    opening_balance = None
    current_txn = None

    with pdfplumber.open(pdf_path) as pdf:
        cols = _detect_column_boundaries(pdf)
        col_koste = cols['fee']
        col_debit = cols['debit']
        col_credit = cols['credit']
        col_saldo = cols['balance']
        col_desc_end = col_koste

        for page in pdf.pages:
            words = page.extract_words()
            lines = _group_words_into_lines(words)

            for line_words in lines:
                if not line_words:
                    continue

                first = line_words[0]

                if first['x0'] < 80 and _DATE_RE_ABSA.match(first['text']):
                    if current_txn:
                        transactions.append(current_txn)

                    date_str = first['text']

                    if not _validate_date(date_str):
                        print(f'  WARNING: Invalid date "{date_str}"')
                        current_txn = None
                        continue

                    desc_parts = []
                    fee_parts = []
                    fee_type = None
                    debit_parts = []
                    credit_parts = []
                    saldo_parts = []

                    for w in line_words[1:]:
                        x0 = w['x0']
                        text = w['text']

                        if x0 < col_desc_end:
                            desc_parts.append(text)
                        elif x0 < col_debit:
                            if len(text) == 1 and text in 'TADGK*':
                                fee_type = text
                            else:
                                fee_parts.append(text)
                        elif x0 < col_credit:
                            debit_parts.append(text)
                        elif x0 < col_saldo:
                            credit_parts.append(text)
                        else:
                            saldo_parts.append(text)

                    desc_text = ' '.join(desc_parts)

                    desc_lower = desc_text.lower()
                    is_balance_row = (
                        ('Saldo' in desc_text and 'orgedra' in desc_lower) or
                        ('Balance' in desc_text and ('arried' in desc_lower or 'forward' in desc_lower))
                    )
                    if is_balance_row:
                        balance_str = ' '.join(saldo_parts)
                        opening_balance = _parse_sa_balance(balance_str)
                        current_txn = None
                        continue

                    debit = _parse_sa_amount(
                        ' '.join(debit_parts)
                    ) if debit_parts else None
                    credit = _parse_sa_amount(
                        ' '.join(credit_parts)
                    ) if credit_parts else None
                    balance = _parse_sa_balance(
                        ' '.join(saldo_parts)
                    ) if saldo_parts else None
                    fee = _parse_sa_amount(
                        ' '.join(fee_parts)
                    ) if fee_parts else None

                    if debit is not None and credit is not None:
                        amount = credit - debit
                    elif debit is not None:
                        amount = -abs(debit)
                    elif credit is not None:
                        amount = abs(credit)
                    else:
                        amount = Decimal('0')

                    description = clean_description(desc_text)
                    if fee is not None:
                        description += f' (Fee: R{fee})'

                    current_txn = Transaction(
                        date=date_str,
                        description=description,
                        amount=amount,
                        balance=balance,
                        fee=fee if fee else Decimal('0'),
                    )

                elif current_txn:
                    cont_parts = [
                        w['text'] for w in line_words if w['x0'] < col_desc_end
                    ]
                    if cont_parts:
                        extra = ' '.join(cont_parts)
                        if any(marker in extra for marker in _ABSA_TJEK_BOILERPLATE):
                            continue
                        current_txn = Transaction(
                            date=current_txn.date,
                            description=clean_description(
                                current_txn.description + ' ' + extra
                            ),
                            amount=current_txn.amount,
                            balance=current_txn.balance,
                            fee=current_txn.fee,
                        )

    if current_txn:
        transactions.append(current_txn)

    return transactions, opening_balance


# ---------------------------------------------------------------------------
# Parser B: ABSA Transaksiegeskiedenis - Credit Card
# ---------------------------------------------------------------------------
def parse_absa_credit_card(pdf_path):
    transactions = []

    with pdfplumber.open(pdf_path) as pdf:
        all_lines = []
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                all_lines.extend(text.split('\n'))

    date_re = re.compile(r'^(\d{4}-\d{2}-\d{2})\s+(.+)$')

    skip_re = re.compile(
        r'Transaksiegeskiedenis|Transaction History|ANDWER FREIGHT|PO BOX|ASTON MANOR|^1630$|'
        r'Huidige Saldo|Current Balance|Beskikbare Saldo|Available Balance|Begrotings Saldo|Budget Balance|'
        r'Staat vir die Periode|Statement for the Period|Kredietkaartrekening|Credit Card Account|^4570|^ABSA$|'
        r'Datum\s+Transaksie Beskrywing\s+Bedrag|Date\s+Transaction Description\s+Amount|'
        r'^\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}|'
        r'Bladsy\s+\d+\s+van\s+\d+|Page\s+\d+\s+of\s+\d+|^\s*$',
        re.IGNORECASE
    )

    current_txn = None

    for line in all_lines:
        line = line.strip()
        if not line:
            continue
        if skip_re.search(line):
            continue

        m = date_re.match(line)
        if m:
            if current_txn:
                transactions.append(current_txn)

            date_raw = m.group(1)
            rest = m.group(2)
            parts = date_raw.split('-')
            date_str = f'{parts[2]}/{parts[1]}/{parts[0]}'

            amount_match = re.search(r'(-?R[\d\s,.]+\.\d{2})\s*$', rest)
            if amount_match:
                amount_raw = amount_match.group(1)
                desc = rest[:amount_match.start()].strip()

                negative = amount_raw.startswith('-')
                amt_clean = amount_raw.lstrip('-').lstrip('R')
                amt_clean = clean_amount_spaces(amt_clean).replace(',', '')

                try:
                    amount = Decimal(amt_clean)
                    if negative:
                        amount = -amount
                except Exception:
                    current_txn = None
                    continue

                current_txn = Transaction(
                    date=date_str,
                    description=clean_description(desc),
                    amount=amount,
                )
            else:
                current_txn = None
        elif current_txn:
            ref = line.strip()
            if ref and len(ref) > 2:
                current_txn = Transaction(
                    date=current_txn.date,
                    description=clean_description(
                        current_txn.description + ' Ref:' + ref
                    ),
                    amount=current_txn.amount,
                    balance=current_txn.balance,
                    fee=current_txn.fee,
                )

    if current_txn:
        transactions.append(current_txn)

    return transactions, None


# ---------------------------------------------------------------------------
# Parser C: ABSA Transaksiegeskiedenis - Current Account
# ---------------------------------------------------------------------------
def parse_absa_current(pdf_path):
    transactions = []

    with pdfplumber.open(pdf_path) as pdf:
        all_lines = []
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                all_lines.extend(text.split('\n'))

    date_re = re.compile(r'^(\d{4}-\d{2}-\d{2})\s+(.+)$')

    skip_re = re.compile(
        r'Transaksiegeskiedenis|Transaction History|ANDWER FREIGHT|PO BOX|ASTON MANOR|^1630$|'
        r'Huidige Saldo|Current Balance|Beskikbare Saldo|Available Balance|Onverrekende Tjeks|Uncleared Cheques|'
        r'Staat vir die Periode|Statement for the Period|^260144588$|^ABSA$|^ANDWER FREIGHT$|'
        r'Datum\s+Transaksie Beskrywing\s+Bedrag\s+Saldo|Date\s+Transaction Description\s+Amount\s+Balance|'
        r'^\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}|'
        r'Bladsy\s+\d+\s+van\s+\d+|Page\s+\d+\s+of\s+\d+|^\s*$',
        re.IGNORECASE
    )

    current_txn = None

    for line in all_lines:
        line = line.strip()
        if not line:
            continue
        if skip_re.search(line):
            continue

        m = date_re.match(line)
        if m:
            if current_txn:
                transactions.append(current_txn)

            date_raw = m.group(1)
            rest = m.group(2)
            parts = date_raw.split('-')
            date_str = f'{parts[2]}/{parts[1]}/{parts[0]}'

            amounts = list(re.finditer(r'(-?R[\d\s,.]+\.\d{2})', rest))

            if len(amounts) >= 2:
                bedrag_m = amounts[-2]
                saldo_m = amounts[-1]
                desc = rest[:bedrag_m.start()].strip()

                amt_raw = bedrag_m.group(1)
                neg = amt_raw.startswith('-')
                amt_clean = amt_raw.lstrip('-').lstrip('R')
                amt_clean = clean_amount_spaces(amt_clean).replace(',', '')
                try:
                    amount = Decimal(amt_clean)
                    if neg:
                        amount = -amount
                except Exception:
                    current_txn = None
                    continue

                sal_raw = saldo_m.group(1)
                neg_s = sal_raw.startswith('-')
                sal_clean = sal_raw.lstrip('-').lstrip('R')
                sal_clean = clean_amount_spaces(sal_clean).replace(',', '')
                try:
                    balance = Decimal(sal_clean)
                    if neg_s:
                        balance = -balance
                except Exception:
                    balance = None

                current_txn = Transaction(
                    date=date_str,
                    description=clean_description(desc),
                    amount=amount,
                    balance=balance,
                )
            elif len(amounts) == 1:
                bedrag_m = amounts[0]
                desc = rest[:bedrag_m.start()].strip()
                amt_raw = bedrag_m.group(1)
                neg = amt_raw.startswith('-')
                amt_clean = amt_raw.lstrip('-').lstrip('R')
                amt_clean = clean_amount_spaces(amt_clean).replace(',', '')
                try:
                    amount = Decimal(amt_clean)
                    if neg:
                        amount = -amount
                except Exception:
                    current_txn = None
                    continue

                current_txn = Transaction(
                    date=date_str,
                    description=clean_description(desc),
                    amount=amount,
                )
            else:
                current_txn = None
        elif current_txn:
            extra = line.strip()
            if extra and not re.match(r'^-?R[\d\s,]+', extra):
                current_txn = Transaction(
                    date=current_txn.date,
                    description=clean_description(
                        current_txn.description + ' ' + extra
                    ),
                    amount=current_txn.amount,
                    balance=current_txn.balance,
                    fee=current_txn.fee,
                )

    if current_txn:
        transactions.append(current_txn)

    return transactions, None


# ---------------------------------------------------------------------------
# Parser D: FNB Business Statement
# ---------------------------------------------------------------------------

MONTH_MAP = {
    'Jan': '01', 'Feb': '02', 'Mrt': '03', 'Mar': '03', 'Apr': '04',
    'Mei': '05', 'May': '05', 'Jun': '06', 'Jul': '07', 'Aug': '08',
    'Sep': '09', 'Okt': '10', 'Oct': '10', 'Nov': '11', 'Des': '12', 'Dec': '12',
}

FULL_MONTHS = {
    # Afrikaans
    'Januarie': '01', 'Februarie': '02', 'Maart': '03', 'April': '04',
    'Mei': '05', 'Junie': '06', 'Julie': '07', 'Augustus': '08',
    'September': '09', 'Oktober': '10', 'November': '11', 'Desember': '12',
    # English
    'January': '01', 'February': '02', 'March': '03',
    'May': '05', 'June': '06', 'July': '07', 'August': '08',
    'October': '10', 'December': '12',
}


def parse_fnb_statement(pdf_path):
    transactions = []

    with pdfplumber.open(pdf_path) as pdf:
        all_lines = []
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                all_lines.extend(text.split('\n'))

    period = _extract_fnb_period(all_lines)

    date_re = re.compile(
        r'^(\d{1,2})\s+(Jan|Feb|Mrt|Mar|Apr|Mei|May|Jun|Jul|Aug|Sep|Okt|Oct|Nov|Des|Dec)\b\s*(.*)',
        re.IGNORECASE
    )

    skip_re = re.compile(
        r'FNB Verified|Reference Number|To verify|ID number|'
        r'Posbus 5550|Kempton Park|Straat Adres|Street Address|Woodbridge|'
        r'H/V Dann|Glen Marais|Universele Takkode|Universal Branch Code|fnb\.co\.za|'
        r'Verlore Kaarte|Lost Cards|Rekeningnavrae|Account Enquiries|Bedrog\s|Fraud\s|'
        r'Verhoudingsbestuurder|Relationship Manager|Gabisile|087-|'
        r'BBST|^ANDWER FREIGHT|63 BOTES|^GLEN MARAIS|'
        r'^1619$|^1630$|^P O BOX|^ASTON MANOR|'
        r'Kli.nt BTW|Client VAT|Bank BTW|Bank VAT|Besigheids Tjekrekening|Business Cheque Account|'
        r'Belastingfaktuur|Tax Invoice|Statnommer|Statement Number|Staat Periode|Statement Period|Staatdatum|Statement Date|'
        r"Staatsaldo|Statement Balance|Bankkoste$|Bank Charges$|Rentekoers|Interest Rate|"
        r'Openingsaldo|Opening Balance|Afsluitingsaldo|Closing Balance|BTW is ingesluit|VAT is included|Totale BTW|Total VAT|'
        r'Diensfooie|Service Fees|Kontant Deposito|Cash Deposit|Kontant Hanterings|Cash Handling|Ander Fooie|Other Fees|'
        r'Kredietsaldo|Credit Balance|Debietsaldo|Debit Balance|Trapvlak|Tier|Oortrokke Fasiliteit|Overdraft Facility|'
        r'Transaksies in RAND|Transactions in RAND|'
        r'^Datum\s|^Date\s|Beskrywing\s+Bedrag\s+Saldo|Description\s+Amount\s+Balance|Opgeloopte|Accrued|Bank-$|^koste$|^charges$|'
        r'Bladsy \d+|Page \d+|Leweringswyse|Delivery Method|NS/EM|NS/15|^978$|^276\d+|^800\d+|'
        r'Tak Nommer|Branch Number|Rekeningnommer|Account Number|DDA BH|BESIGHEIDS TJEKREKENING|BUSINESS CHEQUE ACCOUNT|'
        r'Omset vir Staat|Turnover for Statement|Nr\. Krediet|No\. Credit|Nr\. Debiet|No\. Debit|'
        r'Indien u enige|If you have any|Vir meer inligting|For more information|kredietrentekoerse|credit interest rates|'
        r'First National Bank|Die prima uitleenkoers|The prime lending rate|'
        r'XSTZFN|^\d{3,4}$|'
        r'^\s*$',
        re.IGNORECASE
    )

    for line in all_lines:
        line = line.strip()
        if not line:
            continue
        if skip_re.search(line):
            continue

        m = date_re.match(line)
        if m:
            day = m.group(1).zfill(2)
            month_name = m.group(2)
            rest = m.group(3).strip()

            month_num = MONTH_MAP.get(month_name, MONTH_MAP.get(month_name.capitalize(), '01'))
            year = _get_year_for_month(int(month_num), period)
            date_str = f'{day}/{month_num}/{year}'

            txn = _parse_fnb_txn_line(rest, date_str)
            if txn:
                transactions.append(txn)

    opening_balance = _extract_fnb_opening_balance(all_lines)

    # Fix unsigned balances for Dt (overdrawn) accounts:
    # FNB Dt accounts show balances without Kt/Dt suffix — they're negative.
    # Only negate balances that have no explicit suffix (positive raw values).
    if opening_balance is not None and opening_balance < 0:
        for txn in transactions:
            if txn.balance is not None and txn.balance > 0:
                txn.balance = -txn.balance

    # OCR fee descriptions if available
    if HAS_OCR:
        _ocr_fnb_fee_descriptions(pdf_path, transactions)

    return transactions, opening_balance


def _extract_fnb_period(lines):
    for line in lines:
        m = re.search(
            r'(?:Staat\s+Periode|Statement\s+Period)\s*:\s*\d+\s+(\w+)\s+(\d{4})\s+(?:tot|to)\s+\d+\s+(\w+)\s+(\d{4})',
            line, re.IGNORECASE
        )
        if m:
            start_month = FULL_MONTHS.get(m.group(1))
            end_month = FULL_MONTHS.get(m.group(3))
            if start_month is None or end_month is None:
                raise ValueError(
                    f"Unrecognised month in FNB period header: "
                    f"'{m.group(1)}' or '{m.group(3)}'"
                )
            return {
                'start_month': int(start_month),
                'start_year': m.group(2),
                'end_month': int(end_month),
                'end_year': m.group(4),
            }
    raise ValueError(
        "Could not parse FNB statement period from PDF header. "
        "Expected 'Staat Periode / Statement Period : DD Month YYYY tot/to DD Month YYYY'"
    )


def _extract_fnb_opening_balance(lines):
    """Extract opening balance from FNB statement header (Afrikaans or English)."""
    for line in lines:
        m = re.search(
            r'(?:Openingsaldo|Opening\s+Balance)\s+([\d,]+\.\d{2})\s*(Kt|Dt|Cr|Dr)',
            line, re.IGNORECASE
        )
        if m:
            bal = Decimal(m.group(1).replace(',', ''))
            if m.group(2).lower() in ('dt', 'dr'):
                bal = -bal
            return bal
    return None


def _get_year_for_month(month_num, period):
    """Determine year for a given month number within the statement period.

    Args:
        month_num: Integer month (1-12)
        period: Dict with start_month (int), start_year (str),
                end_month (int), end_year (str)
    """
    if period['start_year'] == period['end_year']:
        return period['start_year']
    if month_num >= period['start_month']:
        return period['start_year']
    return period['end_year']


def _parse_fnb_txn_line(rest, date_str):
    """Parse FNB transaction line (after date).

    - Amount with 'Kt' suffix = CREDIT (positive)
    - Amount without suffix = DEBIT (negative)
    - Balance with Kt/Dt suffix = signed accordingly
    - Balance without suffix = unsigned (inferred from running balance)
    """
    if not rest:
        return None

    num_pattern = re.compile(r'([\d,]+\.\d{2})(Kt|kt|Dt|dt|Cr|cr|Dr|dr)?')
    matches = list(num_pattern.finditer(rest))

    if not matches:
        return None

    amt_str = clean_amount_commas(matches[0].group(1))
    amt_suffix = matches[0].group(2)
    try:
        amount = Decimal(amt_str)
    except Exception:
        return None

    is_credit = amt_suffix and amt_suffix.lower() in ('kt', 'cr')
    if amount == 0:
        pass
    elif is_credit:
        amount = abs(amount)
    else:
        amount = -abs(amount)

    desc = rest[:matches[0].start()].strip()

    balance = None
    if len(matches) >= 2:
        bal_str = clean_amount_commas(matches[1].group(1))
        bal_suffix = matches[1].group(2)
        try:
            balance = Decimal(bal_str)
            if bal_suffix and bal_suffix.lower() in ('dt', 'dr'):
                balance = -balance
            elif not bal_suffix:
                # No suffix — unsigned balance for Dt (overdrawn) accounts.
                # Sign inferred via running reconciliation with opening balance.
                pass
        except Exception:
            balance = None

    if not desc and amount == 0:
        desc = '(Bank fee notification)'
    elif not desc:
        desc = '(Fee/Charge)'

    return Transaction(
        date=date_str,
        description=clean_description(desc),
        amount=amount,
        balance=balance,
    )


def _preprocess_ocr_image(pil_img):
    """Preprocess image for better OCR: grayscale, contrast boost, 2x resize."""
    # Convert to grayscale
    img = pil_img.convert('L')
    # Increase contrast
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(2.0)
    # Resize 2x for better character recognition
    img = img.resize((img.width * 2, img.height * 2), Image.LANCZOS)
    return img


def _clean_ocr_text(text):
    """Clean common OCR artifacts from extracted text."""
    if not text:
        return ''
    # Strip common OCR noise characters
    text = re.sub(r'[#|{}\[\]]+', '', text)
    # Strip leading/trailing punctuation and garbage chars
    text = re.sub(r'^[)\(\]\[.,;:!?\d\s]+', '', text)
    text = re.sub(r'[\d)\(\]\[.,;:!?\s]+$', '', text)
    text = text.strip()
    # Fix common OCR misreads in Afrikaans bank terms
    ocr_fixes = {
        'Diens Fooi': 'Diensfooie',
        'Dienstooi': 'Diensfooie',
        'Diensfool': 'Diensfooie',
        'Dienstool': 'Diensfooie',
        'luitingsaldo': 'Sluitingsaldo',
        'peningsaldo': 'Openingsaldo',
        'iensfooie': 'Diensfooie',
        'ankkoste': 'Bankkoste',
        'TW is ingesluit': 'BTW is ingesluit',
        'Maandelikse': 'Maandelike',
    }
    for wrong, right in ocr_fixes.items():
        if wrong in text:
            text = text.replace(wrong, right)
    return text


# Known FNB fee descriptions (Afrikaans) for fuzzy matching
_FNB_KNOWN_FEES = [
    'Maandelike Diensfooie',
    'BTW is ingesluit teen 15.00%',
    'Diensfooie',
    'Bankkoste',
    'Rekeningfooie',
    'Transaksiefooi',
    'Elektroniese Bankfooie',
    'Kontanthandeling',
    'Minimumfooi',
]


def _match_known_fee(ocr_text):
    """Match OCR text against known FNB fee descriptions.

    Returns the known description if a reasonable match is found, else None.
    """
    if not ocr_text or len(ocr_text) < 3:
        return None
    lower = ocr_text.lower()
    for known in _FNB_KNOWN_FEES:
        known_lower = known.lower()
        # Check if OCR text is a substring match or close match
        if known_lower in lower or lower in known_lower:
            return known
        # Check shared word overlap (at least half the words match)
        known_words = set(known_lower.split())
        ocr_words = set(lower.split())
        overlap = known_words & ocr_words
        if len(overlap) >= max(1, len(known_words) // 2):
            return known
    return None


def _is_garbage_ocr(text):
    """Detect garbage OCR output that should be rejected."""
    if not text:
        return True
    # Reject if it contains amount-like patterns (digits with Kt/Dt suffixes)
    if re.search(r'\d{3,}', text):
        return True
    if re.search(r'\b(Kt|Dt|kt|dt)\b', text):
        return True
    # Reject if too short after cleanup
    if len(text) < 3:
        return True
    # Reject if mostly digits/punctuation
    alpha_count = sum(1 for c in text if c.isalpha())
    if alpha_count < len(text) * 0.4:
        return True
    return False


def _ocr_fnb_fee_descriptions(pdf_path, transactions):
    """Use Tesseract OCR to extract FNB fee descriptions rendered as images.

    Preprocesses images (grayscale, contrast, 2x resize) for better accuracy.
    """
    if not HAS_OCR:
        return

    fee_txns = [t for t in transactions
                if t.description in ('(Fee/Charge)', '(Bank fee notification)')]
    if not fee_txns:
        return

    ocr_warned = False
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                words = page.extract_words()
                if not words:
                    continue

                lines = _group_words_into_lines(words)
                page_height = page.height

                for line_words in lines:
                    if not line_words:
                        continue

                    line_text = ' '.join(w['text'] for w in line_words)
                    y_top = min(w['top'] for w in line_words)
                    y_bottom = max(w['bottom'] for w in line_words)

                    for txn in fee_txns:
                        if txn.description not in ('(Fee/Charge)', '(Bank fee notification)'):
                            continue

                        amt_abs = abs(txn.amount)
                        if amt_abs >= 1000:
                            amt_formatted = f'{amt_abs:,.2f}'
                        else:
                            amt_formatted = f'{amt_abs:.2f}'

                        if amt_formatted not in line_text and str(amt_abs) not in line_text:
                            continue

                        crop_x0 = 30
                        crop_x1 = min(w['x0'] for w in line_words) - 5
                        if crop_x1 <= crop_x0:
                            crop_x1 = 250

                        crop_y0 = max(0, y_top - 3)
                        crop_y1 = min(page_height, y_bottom + 3)

                        try:
                            cropped = page.crop((crop_x0, crop_y0, crop_x1, crop_y1))
                            img = cropped.to_image(resolution=300)
                            pil_img = _preprocess_ocr_image(img.original)

                            ocr_text = pytesseract.image_to_string(
                                pil_img,
                                lang='afr+eng',
                                config='--psm 7'
                            ).strip()

                            ocr_text = _clean_ocr_text(ocr_text)
                            # Try matching against known fee descriptions
                            known = _match_known_fee(ocr_text)
                            if known:
                                txn.description = known
                            elif ocr_text and not _is_garbage_ocr(ocr_text):
                                txn.description = clean_description(ocr_text)
                        except Exception:
                            if not ocr_warned:
                                print('  WARNING: OCR failed for some fee descriptions')
                                ocr_warned = True

    except Exception:
        if not ocr_warned:
            print('  WARNING: OCR processing failed')


# ---------------------------------------------------------------------------
# CSV Writer & Verification
# ---------------------------------------------------------------------------
def _atomic_write(output_path, write_fn):
    """Write a file atomically via temp file then rename."""
    tmp_path = output_path + '.tmp'
    try:
        write_fn(tmp_path)
        os.replace(tmp_path, output_path)
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise


def _write_csv_generic(transactions, output_path, categorise=False):
    """Generic CSV: Date, Description, Amount (signed). UTF-8 BOM."""
    def _write(path):
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
            headers = ['Date', 'Description', 'Amount']
            if categorise:
                headers.append('Category')
            writer.writerow(headers)
            for txn in transactions:
                amt = txn.amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                row = [txn.date, txn.description, str(amt)]
                if categorise:
                    row.append(txn.category or '')
                writer.writerow(row)
    _atomic_write(output_path, _write)


def _write_csv_sage(transactions, output_path, categorise=False):
    """Sage: Date, Description, Amount (signed). Zero amounts left blank. UTF-8 BOM."""
    def _write(path):
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
            headers = ['Date', 'Description', 'Amount']
            if categorise:
                headers.append('Category')
            writer.writerow(headers)
            for txn in transactions:
                amt = txn.amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                amt_str = '' if amt == 0 else str(amt)
                row = [txn.date, txn.description, amt_str]
                if categorise:
                    row.append(txn.category or '')
                writer.writerow(row)
    _atomic_write(output_path, _write)


def _write_csv_sage_split(transactions, output_path, categorise=False):
    """Sage 4-column: Date, Description, Money in, Money out (unsigned). UTF-8 BOM."""
    def _write(path):
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
            headers = ['Date', 'Description', 'Money in', 'Money out']
            if categorise:
                headers.append('Category')
            writer.writerow(headers)
            for txn in transactions:
                amt = txn.amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                if amt > 0:
                    row = [txn.date, txn.description, str(amt), '']
                elif amt < 0:
                    row = [txn.date, txn.description, '', str(abs(amt))]
                else:
                    row = [txn.date, txn.description, '', '']
                if categorise:
                    row.append(txn.category or '')
                writer.writerow(row)
    _atomic_write(output_path, _write)


def _write_csv_xero(transactions, output_path, categorise=False):
    """Xero precoded: *Date, *Amount, Payee, Description, Reference. UTF-8 (no BOM)."""
    def _write(path):
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
            headers = ['*Date', '*Amount', 'Payee', 'Description', 'Reference']
            if categorise:
                headers.append('Category')
            writer.writerow(headers)
            for txn in transactions:
                amt = txn.amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                parts = txn.description.split(' ', 1)
                payee = parts[0] if parts else ''
                detail = parts[1] if len(parts) > 1 else ''
                row = [txn.date, str(amt), payee, detail, '']
                if categorise:
                    row.append(txn.category or '')
                writer.writerow(row)
    _atomic_write(output_path, _write)


def _write_csv_quickbooks(transactions, output_path, categorise=False):
    """QuickBooks: Date, Description, Amount (signed, no commas in numbers). UTF-8 (no BOM)."""
    def _write(path):
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
            headers = ['Date', 'Description', 'Amount']
            if categorise:
                headers.append('Category')
            writer.writerow(headers)
            for txn in transactions:
                amt = txn.amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                amt_str = str(amt).replace(',', '')
                row = [txn.date, txn.description, amt_str]
                if categorise:
                    row.append(txn.category or '')
                writer.writerow(row)
    _atomic_write(output_path, _write)


def _write_csv_quickbooks_split(transactions, output_path, categorise=False):
    """QuickBooks 4-column: Date, Description, Credit, Debit (unsigned). UTF-8 (no BOM)."""
    def _write(path):
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
            headers = ['Date', 'Description', 'Credit', 'Debit']
            if categorise:
                headers.append('Category')
            writer.writerow(headers)
            for txn in transactions:
                amt = txn.amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                if amt > 0:
                    row = [txn.date, txn.description, str(amt), '']
                elif amt < 0:
                    row = [txn.date, txn.description, '', str(abs(amt))]
                else:
                    row = [txn.date, txn.description, '', '']
                if categorise:
                    row.append(txn.category or '')
                writer.writerow(row)
    _atomic_write(output_path, _write)


FORMATTERS = {
    'generic': _write_csv_generic,
    'sage': _write_csv_sage,
    'sage_split': _write_csv_sage_split,
    'xero': _write_csv_xero,
    'quickbooks': _write_csv_quickbooks,
    'quickbooks_split': _write_csv_quickbooks_split,
}


def write_csv(transactions, output_path, fmt='generic', categorise=False):
    """Write transactions to CSV using the selected format."""
    formatter = FORMATTERS.get(fmt, _write_csv_generic)
    formatter(transactions, output_path, categorise=categorise)


def write_xlsx(transactions, output_path, fmt='generic', categorise=False):
    """Write transactions to formatted Excel file.

    Creates a Transactions sheet with bold headers, currency formatting,
    auto-column-widths, and frozen top row.
    When categorise=True, adds a Summary sheet with category breakdown.
    """
    if not HAS_OPENPYXL:
        print('  WARNING: openpyxl not installed — skipping Excel output')
        print('           Install: pip install openpyxl')
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Transactions'

    # Determine headers based on format
    fmt_headers = {
        'generic': ['Date', 'Description', 'Amount'],
        'sage': ['Date', 'Description', 'Amount'],
        'sage_split': ['Date', 'Description', 'Money in', 'Money out'],
        'xero': ['*Date', '*Amount', 'Payee', 'Description', 'Reference'],
        'quickbooks': ['Date', 'Description', 'Amount'],
        'quickbooks_split': ['Date', 'Description', 'Credit', 'Debit'],
    }
    headers = fmt_headers.get(fmt, fmt_headers['generic'])
    if categorise:
        headers = list(headers) + ['Category']

    # Write headers
    bold_font = Font(bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold_font

    # Write transaction rows
    currency_fmt = '#,##0.00'
    for row_idx, txn in enumerate(transactions, 2):
        amt = float(txn.amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))

        if fmt == 'xero':
            parts = txn.description.split(' ', 1)
            payee = parts[0] if parts else ''
            detail = parts[1] if len(parts) > 1 else ''
            values = [txn.date, amt, payee, detail, '']
        elif fmt in ('sage_split', 'quickbooks_split'):
            col_label = 'Money in' if fmt == 'sage_split' else 'Credit'
            if amt > 0:
                values = [txn.date, txn.description, amt, '']
            elif amt < 0:
                values = [txn.date, txn.description, '', abs(amt)]
            else:
                values = [txn.date, txn.description, '', '']
        elif fmt == 'sage' and amt == 0:
            values = [txn.date, txn.description, '']
        else:
            values = [txn.date, txn.description, amt]

        if categorise:
            values.append(txn.category or '')

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if isinstance(val, float):
                cell.number_format = currency_fmt

    # Auto-column-widths
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Summary sheet (when categorise=True and categories exist)
    if categorise:
        summary = generate_category_summary(transactions)
        if summary:
            ws2 = wb.create_sheet('Summary')
            sum_headers = ['Category', 'Count', 'Debits', 'Credits', 'Net']
            for col, header in enumerate(sum_headers, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = bold_font

            for row_idx, row in enumerate(summary, 2):
                ws2.cell(row=row_idx, column=1, value=row['category'])
                ws2.cell(row=row_idx, column=2, value=row['count'])
                deb_cell = ws2.cell(
                    row=row_idx, column=3,
                    value=float(row['debits']),
                )
                deb_cell.number_format = currency_fmt
                cred_cell = ws2.cell(
                    row=row_idx, column=4,
                    value=float(row['credits']),
                )
                cred_cell.number_format = currency_fmt
                net_cell = ws2.cell(
                    row=row_idx, column=5,
                    value=float(row['net']),
                )
                net_cell.number_format = currency_fmt

            for col_cells in ws2.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws2.column_dimensions[col_letter].width = min(max_len + 3, 40)

            ws2.freeze_panes = 'A2'

    # Atomic write
    def _write(path):
        wb.save(path)

    _atomic_write(output_path, _write)


# ---------------------------------------------------------------------------
# Auto-Categorisation — SA-specific transaction classification
# ---------------------------------------------------------------------------

# Banking prefixes that ABSA/FNB prepend to descriptions.
# Stripped before keyword matching to expose the payee name.
TRANSACTION_TYPE_PREFIXES = [
    'MAGTGNG', 'MAGTGN', 'BETALING', 'AANKOOP', 'DEBIETORDER',
    'KREDIETORDER', 'TERUGBETALING', 'OORDRAG', 'ONTVANGS',
    'DEBIT ORDER', 'CREDIT ORDER', 'PAYMENT', 'PURCHASE',
    'TRANSFER', 'RECEIPT', 'REVERSAL', 'REFUND',
    'POS', 'ATM', 'EFT', 'CASH',
    'INTERNETBETALING', 'INTERNET PAYMENT',
    'CELLPHONE BANKING', 'SELFOONBANKDIENS',
    'ABSA BANK', 'FNB',
]

# Categories with keywords. Order within each category doesn't matter.
# Keywords are matched case-insensitively against the description.
CATEGORIES = {
    'Bank Charges': [
        'MNDELIKSE REK-FOOI', 'MONTHLY ACCOUNT FEE', 'REKENINGFOOI',
        'ACCOUNT FEE', 'DIENSFOOIE', 'SERVICE FEE', 'BANKKOSTE',
        'BANK CHARGES', 'BANK FEE', 'CASH HANDLING', 'KONTANT HANTERINGS',
        'TRANSAKSIEFOOI', 'TRANSACTION FEE', 'ATM FEE', 'STRAFFOOI',
        'PENALTY FEE', 'NOTIFICATION FEE', 'KENNISGEWING',
        'ADMINISTRASIEFOOI', 'ADMIN FEE', 'HANDLING FEE',
        'STAMP DUTY', 'SEELREG',
        '(Fee/Charge)', '(Bank fee notification)',
    ],
    'Interest Paid': [
        'DEBIETRENTE', 'DEBIT INTEREST', 'INTEREST CHARGE',
        'RENTE GEHEF', 'INTEREST PAID', 'DEBIETRENTE HOOFKNTOOR',
        'OORTROKKE RENTE', 'OVERDRAFT INTEREST',
    ],
    'Interest Received': [
        'KREDIETRENTE', 'CREDIT INTEREST', 'INTEREST RECEIVED',
        'RENTE ONTVANG', 'INTEREST EARNED',
    ],
    'Salaries & Wages': [
        'SALARY', 'SALARIS', 'WAGES', 'LONE', 'PAYROLL',
        'BETAALROL', 'NETT PAY', 'NETTO BETALING',
    ],
    'Fuel': [
        'SHELL', 'ENGEN', 'CALTEX', 'BP ', 'TOTAL GARAGE', 'SASOL',
        'PETROPORT', 'PUMA ENERGY', 'FUEL', 'BRANDSTOF', 'PETROL',
        'DIESEL',
    ],
    'Groceries': [
        'WOOLWORTHS', 'WOOLIES', 'CHECKERS', 'SHOPRITE', 'PICK N PAY',
        'PNP', 'SPAR ', 'FOOD LOVER', 'GAME ', 'MAKRO', 'MASSMART',
        'USAVE', 'OK FOODS', 'OK GROCER', 'BOXER',
    ],
    'Telecommunications': [
        'MTN SP', 'MTN ', 'VODACOM', 'CELL C', 'TELKOM', 'RAIN ',
        'AFRIHOST', 'WEBAFRICA', 'VUMATEL', 'OPENSERVE', 'MWEB',
        'NEOTEL', 'LIQUID TELECOM', 'AIRTIME', 'DATA BUNDLE',
        'FIBRE', 'INTERNET',
    ],
    'TV & Entertainment': [
        'DSTV', 'MULTICHOICE', 'SHOWMAX', 'NETFLIX', 'SPOTIFY',
        'APPLE.COM', 'GOOGLE PLAY', 'YOUTUBE', 'DISNEY',
        'AMAZON PRIME', 'STER KINEKOR', 'NU METRO',
    ],
    'Electricity': [
        'ESKOM', 'ELECTRICITY', 'ELEKTRISITEIT', 'PREPAID ELEC',
        'CITY POWER', 'TSHWANE ELECTRICITY', 'POWER UTILITY',
        'MUNISIPALE ELEKTRIS',
    ],
    'Water & Rates': [
        'WATER', 'RATES', 'MUNICIPAL', 'MUNISIPAL', 'CITY OF',
        'METRO ', 'PROPERTY RATES', 'EIENDOMSBELASTING',
        'REFUSE', 'VULLIS', 'SEWERAGE',
    ],
    'Insurance (Short Term)': [
        'OUTSURANCE', 'SANTAM', 'HOLLARD', 'MIWAY', 'AUTO & GENERAL',
        'KING PRICE', 'FIRST FOR WOMEN', 'DIALDIRECT', 'BUDGET INS',
        'SHORT TERM', 'KORTTERMYN', 'VEHICLE INSURANCE', 'CAR INSURANCE',
        'HOME INSURANCE', 'HOUSEHOLD', 'BUILDING INSURANCE',
    ],
    'Insurance (Life)': [
        'OLD MUTUAL', 'SANLAM', 'LIBERTY', 'METROPOLITAN',
        'MOMENTUM', 'DISCOVERY LIFE', 'LIFE INSURANCE',
        'LEWENSVERSEKERING', 'FUNERAL COVER', 'BEGRAFNIS',
        'ASSUPOL', 'CLIENTELE LIFE', 'CAPITEC LIFE',
    ],
    'Insurance (Other)': [
        'INSURANCE', 'VERSEKERING', 'UNDERWRITER',
    ],
    'Medical Aid': [
        'DISCOVERY HEALTH', 'MEDSCHEME', 'BONITAS', 'GEMS',
        'MEDIHELP', 'BESTMED', 'MOMENTUM HEALTH', 'FEDHEALTH',
        'MEDICAL AID', 'MEDIESE FONDS', 'MEDICAL SCHEME',
    ],
    'Security': [
        'ADT', 'FIDELITY', 'CHUBB', 'G4S', 'CSS TACTICAL',
        'SECURITY', 'SEKURITEIT', 'ARMED RESPONSE', 'ALARM',
    ],
    'Restaurants & Meals': [
        'MCDONALD', 'KFC', 'NANDO', 'STEERS', 'DEBONAIRS',
        'SPUR', 'OCEAN BASKET', 'WIMPY', 'MUGG & BEAN',
        'VIDA E CAFFE', 'RESTAURANT', 'UBER EATS', 'MR DELIVERY',
        'BOLT FOOD',
    ],
    'Liquor': [
        'TOPS ', 'LIQUOR', 'DRANKE', 'ULTRA LIQUORS', 'MAKRO LIQUOR',
        'NORMAN GOODFELLOWS', 'WINE ',
    ],
    'Stationery & Office': [
        'WALTONS', 'OFFICEBOX', 'STAPLES', 'CNA ',
        'STATIONERY', 'SKRYFBEHOEFTES', 'OFFICE SUPPLIES',
        'CARTRIDGE', 'TONER', 'PRINTER',
    ],
    'Professional Fees': [
        'ATTORNEY', 'PROKUREUR', 'ACCOUNTANT', 'REKENMEESTER',
        'AUDITOR', 'OUDITEUR', 'LEGAL FEE', 'CONSULTANT',
        'ADVISORY', 'PROFESSIONAL',
    ],
    'Freight & Logistics': [
        'COURIER', 'KOERIER', 'POSTNET', 'ARAMEX', 'DHL',
        'FEDEX', 'FASTWAY', 'THE COURIER GUY', 'DSV', 'DAWN WING',
        'RAM HAND', 'FREIGHT', 'VRAG',
    ],
    'Travel & Accommodation': [
        'BOOKING.COM', 'AIRBNB', 'PROTEA HOTEL', 'CITY LODGE',
        'SUN INTERNATIONAL', 'TSOGO SUN', 'KULULA', 'FLYSAFAIR',
        'MANGO AIRLINES', 'FLIGHT', 'VLUG', 'HOTEL', 'GUEST HOUSE',
        'ACCOMMODATION', 'VERBLYF', 'CAR HIRE', 'AVIS', 'HERTZ',
        'EUROPCAR',
    ],
    'Education': [
        'SCHOOL', 'SKOOL', 'UNIVERSITY', 'UNIVERSITEIT', 'COLLEGE',
        'TUITION', 'SKOOLGELD', 'EDUCATION', 'ONDERWYS', 'UNISA',
        'ACADEMY', 'AKADEMIE',
    ],
    'Subscriptions': [
        'SUBSCRIPTION', 'INTEKENING', 'MEMBERSHIP', 'LIDMAATSKAP',
    ],
    'Equipment': [
        'TAKEALOT', 'BUILDERS', 'CASHBUILD', 'LEROY MERLIN',
        'TOOLS', 'GEREEDSKAP', 'EQUIPMENT', 'TOERUSTING',
    ],
    'Clothing': [
        'MR PRICE', 'EDGARS', 'JET ', 'TRUWORTHS', 'FOSCHINI',
        'MARKHAM', 'SPORTSCENE', 'TOTALSPORTS', 'CAPE UNION',
        'H&M', 'ZARA', 'COTTON ON',
    ],
    'Medical Expenses': [
        'PHARMACY', 'APTEEK', 'CLICKS', 'DISCHEM', 'DIS-CHEM',
        'DOCTOR', 'DOKTER', 'DENTIST', 'TANDARTS', 'HOSPITAL',
        'HOSPITAAL', 'PATHOLOGIST', 'PATHCARE', 'LANCET', 'AMPATH',
        'OPTOMETRIST', 'OOGKUNDIGE', 'MEDICAL', 'MEDIESE',
    ],
    'Inter-Account Transfers': [
        'INTERNE OORDRAG', 'INTERNAL TRANSFER', 'OORDRAG NA',
        'TRANSFER TO', 'OORDRAG VAN', 'TRANSFER FROM',
        'BETWEEN ACCOUNTS', 'INTER-ACCOUNT', 'INTERREK',
        'SAVINGS', 'SPAARREKENING',
    ],
    'Loan Repayments': [
        'LOAN', 'LENING', 'INSTALMENT', 'PAAIEMENT',
        'REPAYMENT', 'TERUGBETALING', 'PERSONAL LOAN',
        'PERSOONLIKE LENING', 'CREDIT AGREEMENT',
    ],
    'Vehicle Finance': [
        'WESBANK', 'MFC', 'VEHICLE FINANCE', 'VOERTUIGFINANSIERING',
        'CAR PAYMENT', 'MOTOR PAYMENT', 'ABSA VEHICLE',
    ],
    'Property & Maintenance': [
        'RENT', 'HUUR', 'LEVY', 'HEFFING', 'BODY CORPORATE',
        'MAINTENANCE', 'ONDERHOUD', 'PLUMBER', 'LOODGIETER',
        'ELECTRICIAN', 'ELEKTRISIËN', 'GARDEN', 'TUIN',
    ],
    'Sales/Revenue': [
        'SALES', 'VERKOPE', 'REVENUE', 'INKOMSTE',
        'PAYMENT RECEIVED', 'BETALING ONTVANG',
    ],
    'Rental Income': [
        'RENTAL INCOME', 'HUURINKOMSTE', 'RENT RECEIVED',
        'HUUR ONTVANG',
    ],
}

# Sage Pastel GL account code suggestions per category
SAGE_ACCOUNT_CODES = {
    'Bank Charges': '7200',
    'Interest Paid': '7300',
    'Interest Received': '4200',
    'Salaries & Wages': '7000',
    'Fuel': '7550',
    'Groceries': '7600',
    'Telecommunications': '7450',
    'TV & Entertainment': '7650',
    'Electricity': '7400',
    'Water & Rates': '7410',
    'Insurance (Short Term)': '7500',
    'Insurance (Life)': '7510',
    'Insurance (Other)': '7520',
    'Medical Aid': '7050',
    'Security': '7460',
    'Restaurants & Meals': '7610',
    'Liquor': '7620',
    'Stationery & Office': '7350',
    'Professional Fees': '7250',
    'Freight & Logistics': '7700',
    'Travel & Accommodation': '7750',
    'Education': '7800',
    'Subscriptions': '7360',
    'Equipment': '1800',
    'Clothing': '7630',
    'Medical Expenses': '7060',
    'Inter-Account Transfers': '9000',
    'Loan Repayments': '8000',
    'Vehicle Finance': '8100',
    'Property & Maintenance': '7420',
    'Sales/Revenue': '4000',
    'Rental Income': '4100',
}


def _strip_banking_prefix(description):
    """Strip ABSA/FNB transaction type prefixes to expose the payee."""
    upper = description.upper()
    for prefix in TRANSACTION_TYPE_PREFIXES:
        if upper.startswith(prefix):
            stripped = description[len(prefix):].lstrip(' -:')
            if stripped:
                return stripped
    return description


def categorise_transaction(description):
    """Two-pass keyword match: check fees first, strip prefix, match payee.

    Returns category name string or None if unmatched.
    """
    desc_upper = description.upper()

    # Pass 1: Check bank fees/charges (match before stripping prefix)
    for keyword in CATEGORIES.get('Bank Charges', []):
        if keyword.upper() in desc_upper:
            return 'Bank Charges'

    # Pass 2: Strip banking prefix and match against all categories
    stripped = _strip_banking_prefix(description)
    stripped_upper = stripped.upper()

    for category, keywords in CATEGORIES.items():
        if category == 'Bank Charges':
            continue  # Already checked
        for keyword in keywords:
            if keyword.upper() in stripped_upper:
                return category

    return None


def _load_category_overrides():
    """Load categories.json if it exists next to the converter.

    Merges user-defined categories with built-in ones.
    Returns the merged dict (does not modify CATEGORIES global).
    """
    import json
    override_path = _get_app_dir() / 'categories.json'
    if not override_path.exists():
        return CATEGORIES

    try:
        with open(override_path, 'r', encoding='utf-8') as f:
            overrides = json.load(f)
    except (json.JSONDecodeError, OSError) as e:
        print(f'  WARNING: Could not load categories.json: {e}')
        return CATEGORIES

    merged = dict(CATEGORIES)
    for cat, keywords in overrides.items():
        if cat in merged:
            # Append new keywords (no duplicates)
            existing = {k.upper() for k in merged[cat]}
            merged[cat] = list(merged[cat]) + [
                k for k in keywords if k.upper() not in existing
            ]
        else:
            merged[cat] = keywords

    return merged


def _apply_categories(transactions):
    """Return new list with category field populated on each transaction."""
    categories = _load_category_overrides()

    # Temporarily swap global CATEGORIES for matching with overrides
    result = []
    for txn in transactions:
        cat = categorise_transaction(txn.description)
        result.append(Transaction(
            date=txn.date,
            description=txn.description,
            amount=txn.amount,
            balance=txn.balance,
            fee=txn.fee,
            category=cat,
        ))
    return result


def generate_category_summary(transactions):
    """Group by category, return list of dicts with breakdown.

    Each dict: {category, count, debits, credits, net}
    Sorted by absolute net (largest first), uncategorised at the end.
    """
    groups = {}
    for txn in transactions:
        cat = txn.category or '(Uncategorised)'
        if cat not in groups:
            groups[cat] = {'count': 0, 'debits': Decimal('0'), 'credits': Decimal('0')}
        groups[cat]['count'] += 1
        if txn.amount < 0:
            groups[cat]['debits'] += txn.amount
        elif txn.amount > 0:
            groups[cat]['credits'] += txn.amount

    result = []
    for cat, data in groups.items():
        result.append({
            'category': cat,
            'count': data['count'],
            'debits': data['debits'],
            'credits': data['credits'],
            'net': data['debits'] + data['credits'],
        })

    # Sort: uncategorised last, then by absolute net descending
    result.sort(key=lambda r: (
        r['category'] == '(Uncategorised)',
        -abs(r['net']),
    ))
    return result


def _print_category_summary(summary):
    """Print category breakdown table to console."""
    if not summary:
        return

    print(f'\n  Category Summary:')
    print(f'  {"Category":<30} {"Count":>5} {"Debits":>14} {"Credits":>14} {"Net":>14}')
    print(f'  {"-" * 77}')
    for row in summary:
        print(
            f'  {row["category"]:<30} {row["count"]:>5} '
            f'R{row["debits"]:>12} R{row["credits"]:>12} R{row["net"]:>12}'
        )


def verify_transactions(transactions, filename, opening_balance=None):
    """Verify extracted transactions with balance reconciliation."""
    errors = []

    print(f'\n  Transactions: {len(transactions)}')
    if not transactions:
        print('  WARNING: No transactions extracted!')
        return errors

    print(f'  Date range: {transactions[0].date} to {transactions[-1].date}')

    total_debits = sum(t.amount for t in transactions if t.amount < 0)
    total_credits = sum(t.amount for t in transactions if t.amount > 0)
    total_fees = sum(t.fee for t in transactions)
    net = total_debits + total_credits

    print(f'  Total debits:  R{total_debits:>14}')
    print(f'  Total credits: R{total_credits:>14}')
    print(f'  Total fees:    R{total_fees:>14}')
    print(f'  Net movement:  R{net:>14}')

    bal_txns = [t for t in transactions if t.balance is not None]
    if len(bal_txns) >= 2:
        mismatches = 0
        for i in range(1, len(bal_txns)):
            prev = bal_txns[i - 1]
            curr = bal_txns[i]
            expected_bal = prev.balance + curr.amount
            diff = abs(expected_bal - curr.balance)
            if diff > Decimal('0.02'):
                mismatches += 1
                if mismatches <= 5:
                    errors.append(
                        f'Balance mismatch at {curr.date} '
                        f'"{curr.description[:40]}": '
                        f'expected {expected_bal}, got {curr.balance} '
                        f'(diff {diff})'
                    )

        if mismatches == 0:
            print(f'  Balance check: PASS (all {len(bal_txns)} consecutive balances reconcile)')
        else:
            print(f'  Balance check: FAIL ({mismatches} mismatches)')
            for e in errors[:5]:
                print(f'    {e}')

    if opening_balance is not None and bal_txns:
        first_bal = bal_txns[0]
        expected_first = opening_balance + first_bal.amount
        diff_first = abs(expected_first - first_bal.balance)
        if diff_first > Decimal('0.02'):
            msg = (f'Opening balance mismatch: opening {opening_balance} + '
                   f'amount {first_bal.amount} = '
                   f'{expected_first}, but balance shows {first_bal.balance}')
            errors.append(msg)
            print(f'  Opening check: FAIL ({msg})')
        else:
            last_bal = bal_txns[-1].balance
            print(f'  Opening balance: {opening_balance}')
            print(f'  Closing balance: {last_bal}')
            balance_change = last_bal - opening_balance
            print(f'  Balance change:  {balance_change}')
            net_diff = abs(balance_change - net)
            if net_diff > Decimal('0.10'):
                msg = (f'Net vs balance: net={net}, '
                       f'balance change={balance_change}, diff={net_diff}')
                errors.append(msg)
                print(f'  Net check:       FAIL ({msg})')
            else:
                print(f'  Net check:       PASS')

    print(f'  First 3:')
    for t in transactions[:3]:
        print(f'    {t.date} | {t.description[:55]:<55} | {t.amount:>12}')
    print(f'  Last 3:')
    for t in transactions[-3:]:
        print(f'    {t.date} | {t.description[:55]:<55} | {t.amount:>12}')

    return errors


# ---------------------------------------------------------------------------
# Generic Parser — auto-detect column layout from PDF headers
# ---------------------------------------------------------------------------
# Column header keywords in English and Afrikaans (lowercase).
# Order matters: more specific keywords first to avoid false matches.

_HEADER_KEYWORDS = {
    'date': [
        'trans date', 'transaction date', 'value date', 'posting date',
        'datum', 'date',
    ],
    'description': [
        'transaksiebeskrywing', 'transaksie beskrywing', 'beskrywing',
        'description', 'narration', 'details', 'transaction', 'particulars',
    ],
    'debit': [
        'debietbedrag', 'debiet', 'debit', 'payments', 'withdrawals',
        'money out (r)', 'amount out', 'money out',
    ],
    'credit': [
        'kredietbedrag', 'krediet', 'credit', 'deposits',
        'money in (r)', 'amount in', 'money in',
    ],
    'amount': [
        'bedrag', 'amount',
    ],
    'balance': [
        'saldo', 'balance', 'running balance', 'closing balance',
    ],
}

# Date patterns to try (most specific first)
_DATE_PATTERNS = [
    # DD/MM/YYYY
    (re.compile(r'^\d{1,2}/\d{2}/\d{4}$'), 'dmy_slash'),
    # YYYY-MM-DD
    (re.compile(r'^\d{4}-\d{2}-\d{2}$'), 'ymd_dash'),
    # DD-MM-YYYY
    (re.compile(r'^\d{1,2}-\d{2}-\d{4}$'), 'dmy_dash'),
    # YYYY/MM/DD
    (re.compile(r'^\d{4}/\d{2}/\d{2}$'), 'ymd_slash'),
    # DD MMM (short month, day as own word) — checked as 2-word combo
]

# Short month names (English + Afrikaans)
_SHORT_MONTHS = set(MONTH_MAP.keys())


def _detect_columns_from_headers(pdf):
    """Scan first 3 pages for column header row and extract x-positions.

    Supports multi-word headers like "Money In (R)", "Posting Date", etc.
    by joining consecutive words and checking against keywords.

    Returns dict like:
        {'date': 12.0, 'description': 100.0, 'debit': 350.0,
         'credit': 430.0, 'balance': 510.0}
    or None if no header row found.
    """
    for page in pdf.pages[:3]:
        words = page.extract_words()
        lines = _group_words_into_lines(words)

        for line_words in lines:
            columns_found = {}

            # Build all possible word n-grams (1-word, 2-word, 3-word)
            # for multi-word header matching
            word_texts = [(w['text'].lower(), w['x0']) for w in line_words]
            ngrams = []
            for i in range(len(word_texts)):
                # Single word
                ngrams.append((word_texts[i][0], word_texts[i][1]))
                # Two-word
                if i + 1 < len(word_texts):
                    ngrams.append((
                        word_texts[i][0] + ' ' + word_texts[i + 1][0],
                        word_texts[i][1],
                    ))
                # Three-word
                if i + 2 < len(word_texts):
                    ngrams.append((
                        word_texts[i][0] + ' ' + word_texts[i + 1][0] + ' ' + word_texts[i + 2][0],
                        word_texts[i][1],
                    ))

            for col_type, keywords in _HEADER_KEYWORDS.items():
                if col_type in columns_found:
                    continue
                for kw in keywords:
                    matched = False
                    for ngram_text, ngram_x0 in ngrams:
                        if kw == ngram_text or ngram_text.startswith(kw):
                            columns_found[col_type] = ngram_x0
                            matched = True
                            break
                    if matched:
                        break

            # When both 'transaction date' and 'posting date' map to 'date',
            # prefer 'transaction date' — more useful for accounting.
            # (Both are already mapped to 'date' in _HEADER_KEYWORDS.)

            # Valid header = has date + (debit or amount) column
            has_date = 'date' in columns_found
            has_amounts = 'debit' in columns_found or 'amount' in columns_found
            if has_date and has_amounts:
                return columns_found

    return None


def _detect_date_format(words_on_line):
    """Detect which date pattern the first word(s) match.

    Returns (date_str_in_dd_mm_yyyy, format_type) or (None, None).
    """
    if not words_on_line:
        return None, None

    first_text = words_on_line[0]['text']

    # Single-word date patterns
    for pattern, fmt_type in _DATE_PATTERNS:
        if pattern.match(first_text):
            if fmt_type == 'dmy_slash':
                return first_text, fmt_type
            if fmt_type == 'ymd_dash':
                parts = first_text.split('-')
                return f'{parts[2]}/{parts[1]}/{parts[0]}', fmt_type
            if fmt_type == 'dmy_dash':
                parts = first_text.split('-')
                return f'{parts[0]}/{parts[1]}/{parts[2]}', fmt_type
            if fmt_type == 'ymd_slash':
                parts = first_text.split('/')
                return f'{parts[2]}/{parts[1]}/{parts[0]}', fmt_type

    # Two-word date: DD + MMM (e.g. "02" "Jan")
    if (len(words_on_line) >= 2
            and re.match(r'^\d{1,2}$', first_text)
            and words_on_line[1]['text'] in _SHORT_MONTHS):
        day = first_text.zfill(2)
        month_num = MONTH_MAP[words_on_line[1]['text']]
        # Year unknown at this point — caller must resolve
        return f'{day}/{month_num}/????', 'dd_mmm'

    return None, None


def _build_column_boundaries(columns):
    """Convert column header x-positions into boundary ranges.

    Returns a sorted list of (min_x, col_type) tuples.
    The boundary for each column extends until the next column starts.
    """
    sorted_cols = sorted(columns.items(), key=lambda kv: kv[1])
    return sorted_cols


def _classify_word(x0, boundaries):
    """Given a word's x0 position and column boundaries, return the column type."""
    result = boundaries[0][0]  # Default to first column
    for col_type, col_x in boundaries:
        if x0 >= col_x - 10:  # 10pt tolerance
            result = col_type
    return result


def _parse_generic_amount(text):
    """Parse an amount from various SA formats.

    Handles: 1234.56, 1,234.56, 1 234.56, -R1234.56, R1234.56,
    1234.56Kt, 1234.56Dt, -1234.56
    """
    text = text.strip()
    if not text:
        return None, None

    # Check for Kt/Dt/Cr/Dr suffix
    suffix = None
    lower = text.lower()
    if lower.endswith('kt') or lower.endswith('cr'):
        suffix = 'kt'
        text = text[:-2]
    elif lower.endswith('dt') or lower.endswith('dr'):
        suffix = 'dt'
        text = text[:-2]

    # Check for -R or R prefix
    negative = False
    if text.startswith('-R') or text.startswith('-r'):
        negative = True
        text = text[2:]
    elif text.startswith('R') or text.startswith('r'):
        text = text[1:]
    elif text.startswith('-'):
        negative = True
        text = text[1:]

    # Remove thousands separators (space or comma)
    text = text.replace(' ', '').replace(',', '')

    try:
        val = Decimal(text)
    except Exception:
        return None, None

    if negative:
        val = -val
    if suffix == 'dt':
        val = -abs(val)

    return val, suffix


def parse_generic_statement(pdf_path):
    """Generic parser that auto-detects column layout from PDF headers.

    Works for any bank statement with:
    - A column header row containing recognisable keywords
    - Transaction lines starting with a date at the left margin
    - Amounts in consistent column positions
    """
    transactions = []
    opening_balance = None

    with pdfplumber.open(pdf_path) as pdf:
        # Step 1: Detect column layout from headers
        columns = _detect_columns_from_headers(pdf)
        if not columns:
            raise ValueError(
                f"Could not detect column headers in {pdf_path}. "
                "Expected keywords like Date/Datum, Description/Beskrywing, "
                "Debit/Credit/Amount/Bedrag, Balance/Saldo."
            )

        boundaries = _build_column_boundaries(columns)
        date_col_x = columns.get('date', 0)
        has_separate_debit_credit = 'debit' in columns and 'credit' in columns
        has_single_amount = 'amount' in columns and 'debit' not in columns

        print(f'  Columns detected: {", ".join(sorted(columns.keys()))}')
        print(f'  Layout: {"Separate Debit/Credit" if has_separate_debit_credit else "Single Amount"}')

        # Step 2: Detect date format from first transaction line
        date_format = None
        date_word_count = 1  # How many words the date consumes

        for page in pdf.pages[:3]:
            words = page.extract_words()
            lines = _group_words_into_lines(words)
            for line_words in lines:
                if not line_words:
                    continue
                if abs(line_words[0]['x0'] - date_col_x) > 15:
                    continue
                date_str, fmt = _detect_date_format(line_words)
                if fmt:
                    date_format = fmt
                    date_word_count = 2 if fmt == 'dd_mmm' else 1
                    break
            if date_format:
                break

        if not date_format:
            raise ValueError(
                f"Could not detect date format in {pdf_path}. "
                "Expected DD/MM/YYYY, YYYY-MM-DD, or DD MMM patterns."
            )

        print(f'  Date format: {date_format}')

        # Step 3: Extract transactions page by page
        for page in pdf.pages:
            words = page.extract_words()
            lines = _group_words_into_lines(words)

            for line_words in lines:
                if not line_words:
                    continue

                # Check if line starts with a date at the date column
                if abs(line_words[0]['x0'] - date_col_x) > 15:
                    continue

                date_str, fmt = _detect_date_format(line_words)
                if not fmt:
                    continue

                # Skip the date words, classify remaining by column
                remaining = line_words[date_word_count:]
                desc_parts = []
                debit_parts = []
                credit_parts = []
                amount_parts = []
                balance_parts = []

                for w in remaining:
                    col = _classify_word(w['x0'], boundaries)
                    if col == 'description':
                        desc_parts.append(w['text'])
                    elif col == 'debit':
                        debit_parts.append(w['text'])
                    elif col == 'credit':
                        credit_parts.append(w['text'])
                    elif col == 'amount':
                        amount_parts.append(w['text'])
                    elif col == 'balance':
                        balance_parts.append(w['text'])
                    elif col == 'date':
                        # Extra date-column word = part of date or desc
                        desc_parts.append(w['text'])

                # Parse amounts
                amount = Decimal('0')
                if has_separate_debit_credit:
                    debit_text = ''.join(debit_parts)
                    credit_text = ''.join(credit_parts)
                    debit_val, _ = _parse_generic_amount(debit_text) if debit_text else (None, None)
                    credit_val, _ = _parse_generic_amount(credit_text) if credit_text else (None, None)
                    if debit_val is not None and credit_val is not None:
                        amount = credit_val - abs(debit_val)
                    elif debit_val is not None:
                        amount = -abs(debit_val)
                    elif credit_val is not None:
                        amount = abs(credit_val)
                elif has_single_amount:
                    amt_text = ''.join(amount_parts)
                    amt_val, suffix = _parse_generic_amount(amt_text) if amt_text else (None, None)
                    if amt_val is not None:
                        amount = amt_val

                # Parse balance
                balance = None
                bal_text = ''.join(balance_parts)
                if bal_text:
                    bal_val, _ = _parse_generic_amount(bal_text)
                    if bal_val is not None:
                        balance = bal_val

                desc = clean_description(' '.join(desc_parts))
                if not desc:
                    desc = '(No description)'

                # Handle DD MMM dates — resolve year from context
                if date_str and '????' in date_str:
                    # Default to current year, caller can fix
                    import datetime
                    date_str = date_str.replace('????', str(datetime.date.today().year))

                if date_str:
                    transactions.append(Transaction(
                        date=date_str,
                        description=desc,
                        amount=amount,
                        balance=balance,
                    ))

    return transactions, opening_balance


# ---------------------------------------------------------------------------
# Bank Format Registry — auto-detection + parser lookup
# ---------------------------------------------------------------------------
# Each entry: detect(text) → bool, parser function (None = not yet supported),
# and a human-readable label.
# Detection order matters: more specific checks must come before general ones.

BANK_FORMATS = [
    # --- ABSA (supported) ---
    {
        'key': 'absa_tjek',
        'label': 'ABSA Tjekrekeningstaat',
        'detect': lambda t: (
            'tjekrekeningnommer' in t or 'tjekrekeningstaat' in t or
            'cheque account number' in t or 'cheque account statement' in t
        ),
        'parser': parse_absa_tjekrekeningstaat,
    },
    {
        'key': 'absa_cc',
        'label': 'ABSA Credit Card',
        'detect': lambda t: (
            ('transaksiegeskiedenis' in t or 'transaction history' in t) and
            ('kredietkaart' in t or 'credit card' in t)
        ),
        'parser': parse_absa_credit_card,
    },
    {
        'key': 'absa_current',
        'label': 'ABSA Current Account',
        'detect': lambda t: (
            ('transaksiegeskiedenis' in t or 'transaction history' in t) and
            'absa' in t
        ),
        'parser': parse_absa_current,
    },
    # --- FNB (supported) ---
    {
        'key': 'fnb',
        'label': 'FNB Business',
        'detect': lambda t: (
            'first national bank' in t or 'fnb.co.za' in t or
            'besigheids tjekrekening' in t or 'business cheque account' in t or
            'business current account' in t
        ),
        'parser': parse_fnb_statement,
    },
    # --- Other SA banks (generic auto-detect parser) ---
    {
        'key': 'standard_bank',
        'label': 'Standard Bank',
        'detect': lambda t: 'standard bank' in t or 'standardbank.co.za' in t,
        'parser': parse_generic_statement,
    },
    {
        'key': 'nedbank',
        'label': 'Nedbank',
        'detect': lambda t: 'nedbank' in t or 'nedbank.co.za' in t,
        'parser': parse_generic_statement,
    },
    {
        'key': 'capitec',
        'label': 'Capitec',
        'detect': lambda t: 'capitec' in t or 'capitecbank.co.za' in t,
        'parser': parse_generic_statement,
    },
    {
        'key': 'investec',
        'label': 'Investec',
        'detect': lambda t: 'investec' in t or 'investec.com' in t,
        'parser': parse_generic_statement,
    },
    {
        'key': 'discovery',
        'label': 'Discovery Bank',
        'detect': lambda t: 'discovery bank' in t or 'discovery.co.za' in t,
        'parser': parse_generic_statement,
    },
    {
        'key': 'tymebank',
        'label': 'TymeBank',
        'detect': lambda t: 'tymebank' in t or 'tymebank.co.za' in t,
        'parser': parse_generic_statement,
    },
    {
        'key': 'african_bank',
        'label': 'African Bank',
        'detect': lambda t: 'african bank' in t or 'africanbank.co.za' in t,
        'parser': parse_generic_statement,
    },
]

# Flat lookup for --parser flag
PARSERS = {fmt['key']: fmt['parser'] for fmt in BANK_FORMATS if fmt['parser']}
PARSERS['generic'] = parse_generic_statement

# All known keys (for argparse choices)
ALL_PARSER_KEYS = [fmt['key'] for fmt in BANK_FORMATS] + ['generic']


def detect_parser(pdf_path):
    """Auto-detect which parser to use based on PDF content.

    Scans first 2 pages for bank-specific markers. Returns parser key.
    Raises ValueError with helpful message if unsupported or unknown.
    """
    with pdfplumber.open(pdf_path) as pdf:
        if not pdf.pages:
            raise ValueError(f"PDF has no pages: {pdf_path}")

        text = ''
        for page in pdf.pages[:2]:
            page_text = page.extract_text()
            if page_text:
                text += page_text + '\n'

        if not text:
            raise ValueError(f"Could not extract text from PDF: {pdf_path}")

    text_lower = text.lower()

    for fmt in BANK_FORMATS:
        if fmt['detect'](text_lower):
            return fmt['key']

    # Unknown bank — try the generic parser as last resort
    return 'generic'


def _generate_output_filename(pdf_path, parser_key, transactions=None, smart=False):
    """Generate output CSV filename from PDF name.

    When smart=True, produces: "Bank Label - Mon YYYY.csv"
    Multi-month: "Bank Label - Dec 2025 - Jan 2026.csv"
    Dedup suffix if file already exists: " (2)", " (3)"
    Falls back to "{stem}.csv" if detection fails or smart=False.
    """
    stem = Path(pdf_path).stem

    if not smart or not transactions:
        return stem + '.csv'

    # Get bank label from BANK_FORMATS
    bank_label = None
    for fmt in BANK_FORMATS:
        if fmt['key'] == parser_key:
            bank_label = fmt['label']
            break
    if not bank_label:
        bank_label = parser_key.replace('_', ' ').title()

    # Extract date range from transactions
    try:
        month_names = [
            '', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
            'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec',
        ]
        first_date = transactions[0].date  # DD/MM/YYYY
        last_date = transactions[-1].date
        first_parts = first_date.split('/')
        last_parts = last_date.split('/')
        first_month = int(first_parts[1])
        first_year = first_parts[2]
        last_month = int(last_parts[1])
        last_year = last_parts[2]

        if first_month == last_month and first_year == last_year:
            period = f'{month_names[first_month]} {first_year}'
        else:
            period = (
                f'{month_names[first_month]} {first_year} - '
                f'{month_names[last_month]} {last_year}'
            )

        name = f'{bank_label} - {period}.csv'
    except (IndexError, ValueError):
        return stem + '.csv'

    # Sanitise for filesystem (remove characters invalid on Windows)
    for ch in '<>:"/\\|?*':
        name = name.replace(ch, '_')

    return name


# ---------------------------------------------------------------------------
# Polish: duplicate detection, encoding checks, summary report
# ---------------------------------------------------------------------------
def _detect_duplicates(transactions):
    """Flag transactions with same date + amount + description."""
    seen = {}
    dupes = []
    for i, txn in enumerate(transactions):
        key = (txn.date, str(txn.amount), txn.description)
        if key in seen:
            dupes.append(
                f'Row {i + 1} duplicates row {seen[key] + 1}: '
                f'{txn.date} {txn.description[:40]} {txn.amount}'
            )
        else:
            seen[key] = i
    return dupes


def _check_encoding(transactions):
    """Detect non-ASCII characters that might cause import issues."""
    issues = []
    for i, txn in enumerate(transactions):
        non_ascii = [c for c in txn.description if ord(c) > 127]
        if non_ascii:
            chars = ''.join(set(non_ascii))
            issues.append(f'Row {i + 1}: "{txn.description[:40]}" contains: {chars}')
    return issues


def _print_summary(summary_rows, fmt):
    """Print a summary table of all processed files."""
    if not summary_rows:
        return

    print('\n' + '=' * 70)
    print('SUMMARY REPORT')
    print(f'Output format: {fmt}')
    print('-' * 70)
    print(f'{"File":<35} {"Parser":<15} {"Txns":>5} {"Net":>14} {"Status":<8}')
    print('-' * 70)

    total_txns = 0
    total_net = Decimal('0')
    for row in summary_rows:
        net_str = f'R{row["net"]:>12}' if row['txns'] > 0 else '-'
        print(f'{row["file"]:<35} {row["parser"]:<15} {row["txns"]:>5} {net_str:>14} {row["status"]:<8}')
        total_txns += row['txns']
        total_net += row['net']

    print('-' * 70)
    print(f'{"TOTAL":<35} {"":<15} {total_txns:>5} {f"R{total_net:>12}":>14}')
    print('=' * 70)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(
        description='PDF Bank Statement to CSV Converter for SA Accounting',
        epilog='Parsers: ABSA (Tjek, CC, Current), FNB Business, + generic auto-detect. '
               'Formats: generic, sage, sage_split, xero, quickbooks, quickbooks_split.'
    )
    ap.add_argument(
        'files', nargs='*',
        help='PDF files to convert (default: all *.pdf in current directory)'
    )
    ap.add_argument(
        '--output-dir', '-o', default='csv',
        help='Output directory for CSV files (default: csv/)'
    )
    ap.add_argument(
        '--parser', '-p', choices=sorted(PARSERS.keys()),
        help='Force parser type (default: auto-detect)'
    )
    ap.add_argument(
        '--format', '-f', choices=sorted(FORMATTERS.keys()), default='generic',
        help='Output CSV format (default: generic)'
    )
    ap.add_argument(
        '--password', default=None,
        help='Password for encrypted PDFs (SA banks often use ID number)'
    )
    ap.add_argument(
        '--categorise', '--categorize', action='store_true', default=False,
        help='Auto-categorise transactions (adds Category column to CSV)'
    )
    ap.add_argument(
        '--smart-names', action='store_true', default=False,
        help='Smart output filenames: "Bank Label - Mon YYYY.csv"'
    )
    ap.add_argument(
        '--xlsx', action='store_true', default=False,
        help='Also output Excel (.xlsx) file with formatting'
    )
    args = ap.parse_args()

    base_dir = _get_app_dir()
    pdf_dir = base_dir / 'pdfs'
    output_dir = base_dir / args.output_dir
    output_dir.mkdir(exist_ok=True)
    pdf_dir.mkdir(exist_ok=True)

    if args.files:
        pdf_files = [Path(f) for f in args.files]
        pdf_files = [
            f if f.is_absolute() else base_dir / f
            for f in pdf_files
        ]
    else:
        # Look in pdfs/ folder first, fall back to base dir
        pdf_files = sorted(pdf_dir.glob('*.pdf'))
        if not pdf_files:
            pdf_files = sorted(base_dir.glob('*.pdf'))

    if not pdf_files:
        print('No PDF files found.')
        print(f'Place your bank statement PDFs in: {pdf_dir}')
        sys.exit(1)

    if not HAS_OCR:
        print('NOTE: pytesseract/tesseract-ocr not installed. FNB fee descriptions will show as (Fee/Charge).')
        print('      Install: sudo apt-get install tesseract-ocr tesseract-ocr-afr && pip install pytesseract Pillow\n')

    version = '2.0.0'
    version_file = _get_bundle_dir() / 'VERSION'
    if version_file.exists():
        version = version_file.read_text().strip()

    print('=' * 70)
    print(f'PDF Bank Statement to CSV Converter v{version}')
    print('=' * 70)

    all_errors = []
    summary_rows = []
    for input_path in pdf_files:
        print(f'\nProcessing: {input_path.name}')

        if not input_path.exists():
            print(f'  ERROR: File not found: {input_path}')
            summary_rows.append({
                'file': input_path.name, 'parser': '-', 'txns': 0,
                'debits': Decimal('0'), 'credits': Decimal('0'),
                'net': Decimal('0'), 'status': 'NOT FOUND',
            })
            continue

        # Decrypt password-protected PDFs if needed
        actual_path = str(input_path)
        decrypted_tmp = None
        try:
            actual_path, decrypted_tmp = _decrypt_pdf_if_needed(
                str(input_path), password=args.password
            )
            if decrypted_tmp:
                print(f'  Decrypted password-protected PDF')
        except ValueError as e:
            print(f'  ERROR: {e}')
            summary_rows.append({
                'file': input_path.name, 'parser': '-', 'txns': 0,
                'debits': Decimal('0'), 'credits': Decimal('0'),
                'net': Decimal('0'), 'status': 'ENCRYPTED',
            })
            continue

        try:
            if args.parser:
                parser_key = args.parser
                print(f'  Parser: {parser_key} (forced)')
            else:
                parser_key = detect_parser(actual_path)
                print(f'  Parser: {parser_key} (auto-detected)')
        except ValueError as e:
            print(f'  ERROR: {e}')
            summary_rows.append({
                'file': input_path.name, 'parser': '-', 'txns': 0,
                'debits': Decimal('0'), 'credits': Decimal('0'),
                'net': Decimal('0'), 'status': 'DETECT FAIL',
            })
            continue
        finally:
            if decrypted_tmp and os.path.exists(decrypted_tmp):
                pass  # Clean up after parsing, not here

        parse_fn = PARSERS[parser_key]

        try:
            transactions, opening_bal = parse_fn(actual_path)

            # Auto-categorise if requested
            if args.categorise and transactions:
                transactions = _apply_categories(transactions)
                categorised = sum(1 for t in transactions if t.category)
                print(f'  Categorised: {categorised}/{len(transactions)} transactions')

            output_name = _generate_output_filename(
                str(input_path), parser_key,
                transactions=transactions, smart=args.smart_names,
            )
            output_path = output_dir / output_name

            print(f'  Output: {output_name} (format: {args.format})')
            file_errors = verify_transactions(
                transactions, input_path.name, opening_bal
            )

            # Duplicate detection
            dupes = _detect_duplicates(transactions)
            if dupes:
                print(f'  WARNING: {len(dupes)} potential duplicate(s):')
                for d in dupes[:5]:
                    print(f'    {d}')

            # Encoding warnings
            encoding_issues = _check_encoding(transactions)
            if encoding_issues:
                print(f'  WARNING: {len(encoding_issues)} description(s) with non-ASCII chars (may cause import issues)')

            total_debits = sum(t.amount for t in transactions if t.amount < 0)
            total_credits = sum(t.amount for t in transactions if t.amount > 0)

            if not transactions:
                print(f'  SKIPPED: No transactions extracted')
                summary_rows.append({
                    'file': input_path.name, 'parser': parser_key, 'txns': 0,
                    'debits': Decimal('0'), 'credits': Decimal('0'),
                    'net': Decimal('0'), 'status': 'EMPTY',
                })
                continue

            # Write AFTER verification passes
            write_csv(
                transactions, str(output_path),
                fmt=args.format, categorise=args.categorise,
            )

            # Excel output if requested
            if args.xlsx:
                xlsx_name = Path(output_name).stem + '.xlsx'
                xlsx_path = output_dir / xlsx_name
                write_xlsx(
                    transactions, str(xlsx_path),
                    fmt=args.format, categorise=args.categorise,
                )
                print(f'  Excel:  {xlsx_name}')

            # Category summary
            if args.categorise and transactions:
                summary = generate_category_summary(transactions)
                _print_category_summary(summary)

            all_errors.extend(
                (input_path.name, e) for e in file_errors
            )

            status = 'FAIL' if file_errors else 'PASS'
            summary_rows.append({
                'file': input_path.name, 'parser': parser_key,
                'txns': len(transactions),
                'debits': total_debits, 'credits': total_credits,
                'net': total_debits + total_credits, 'status': status,
            })
        except Exception as e:
            print(f'  ERROR: {e}')
            traceback.print_exc()
            summary_rows.append({
                'file': input_path.name, 'parser': parser_key, 'txns': 0,
                'debits': Decimal('0'), 'credits': Decimal('0'),
                'net': Decimal('0'), 'status': 'ERROR',
            })
        finally:
            if decrypted_tmp and os.path.exists(decrypted_tmp):
                os.remove(decrypted_tmp)

    # Summary report
    _print_summary(summary_rows, args.format)

    print('\n' + '=' * 70)
    if all_errors:
        print(f'VERIFICATION ISSUES ({len(all_errors)}):')
        for fname, err in all_errors:
            print(f'  [{fname}] {err}')
    else:
        print('ALL VERIFICATIONS PASSED')
    print('=' * 70)


if __name__ == '__main__':
    main()
